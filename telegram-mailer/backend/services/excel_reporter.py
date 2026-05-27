from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from backend.database import get_campaign_stats_by_group, get_campaign_stats_by_account
from datetime import datetime

async def generate_group_report(campaign_id: int) -> str:
    wb = Workbook()
    ws = wb.active
    ws.title = "Группы"
    headers = ["Название группы", "ID группы", "Ссылка", "Кол-во попыток", "Успешные отправки", "Неуспешные отправки", "% успешных", "% ошибок", "Кол-во аккаунтов"]
    ws.append(headers)
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col).font = Font(bold=True)
    data = await get_campaign_stats_by_group(campaign_id)
    for row in data:
        ws.append(row)
    filename = f"reports/campaign_{campaign_id}_groups_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    wb.save(filename)
    return filename

async def generate_account_report(campaign_id: int) -> str:
    wb = Workbook()
    ws = wb.active
    ws.title = "Аккаунты"
    headers = ["Имя аккаунта", "Телефон", "Кол-во групп", "Кол-во попыток", "Успешные отправки", "Неуспешные отправки", "% успешных", "% ошибок"]
    ws.append(headers)
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col).font = Font(bold=True)
    data = await get_campaign_stats_by_account(campaign_id)
    for row in data:
        ws.append(row)
    filename = f"reports/campaign_{campaign_id}_accounts_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    wb.save(filename)
    return filename